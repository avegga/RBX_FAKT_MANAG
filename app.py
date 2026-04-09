import csv
from copy import deepcopy
import io
import json
import re
from datetime import date, datetime
from pathlib import Path

from flask import Flask, jsonify, request, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook, load_workbook
from sqlalchemy import func


BASE_DIR = Path(__file__).resolve().parent
ALLOWED_COLUMN_TYPES = {"text", "choice", "number", "money", "date", "datetime", "boolean"}
COLUMN_TYPE_LABELS = {
    "text": "текст",
    "choice": "список выбора",
    "number": "число",
    "money": "денежный",
    "date": "дата",
    "datetime": "дата-время",
    "boolean": "логический",
}
ALLOWED_ANALYSIS_CHART_TYPES = {"bar", "line", "pie", "table"}
ALLOWED_ANALYSIS_SOURCES = {"none", "facts", "file"}
ALLOWED_ANALYSIS_AGG_FUNCS = {"count", "sum", "avg", "min", "max"}
DEFAULT_STATE = {
    "left_panel_width": 280,
    "right_panel_visible": True,
    "errors_visible": True,
    "errors_height": 180,
    "last_template_id": None,
    "last_loaded_file": "",
}
DEFAULT_ANALYSIS_USER_STATE = {
    "selected_analysis_type_id": None,
    "left_panel_width": 260,
    "visual_source_kind": "none",
    "table_search": "",
    "table_sort_column": "",
    "table_sort_direction": "asc",
    "draft_state": None,
}
EMPTY_DATASET = {
    "file_name": "",
    "header_row_number": None,
    "columns": [],
    "rows": [],
    "warnings": [],
    "errors": [],
    "partial": False,
    "missing_columns": [],
    "row_count": 0,
    "template": None,
    "next_row_id": 1,
    "can_undo_parse": False,
}
EMPTY_ANALYSIS_DATASET = {
    "source_kind": "none",
    "source_mode": "none",
    "source_label": "",
    "source_file_name": "",
    "attached_at": None,
    "source_status": "empty",
    "columns": [],
    "rows": [],
    "row_count": 0,
    "column_count": 0,
    "warnings": [],
}
CURRENT_DATASET = deepcopy(EMPTY_DATASET)
CURRENT_ANALYSIS_DATASET = deepcopy(EMPTY_ANALYSIS_DATASET)
LAST_PARSE_SNAPSHOT = None
SUPPORTED_SETTING_KEYS = ("output_folder", "instructions_file")
APP_SESSION_STARTED_AT = datetime.utcnow().replace(microsecond=0)

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///rbx_fakt_manag.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)


class Template(db.Model):
    __tablename__ = "templates"

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False, unique=True)
    created_at = db.Column(db.DateTime, server_default=func.now(), nullable=False)
    columns = db.relationship(
        "TemplateColumn",
        backref="template",
        cascade="all, delete-orphan",
        order_by="TemplateColumn.position",
        lazy=True,
    )


class TemplateColumn(db.Model):
    __tablename__ = "template_columns"

    id = db.Column(db.Integer, primary_key=True)
    template_id = db.Column(db.Integer, db.ForeignKey("templates.id"), nullable=False)
    column_name = db.Column(db.String(120), nullable=False)
    column_type = db.Column(db.String(40), nullable=False, default="text")
    position = db.Column(db.Integer, nullable=False, default=0)


class AppSetting(db.Model):
    __tablename__ = "app_settings"

    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(120), nullable=False, unique=True)
    value = db.Column(db.Text, nullable=False, default="")


class AppState(db.Model):
    __tablename__ = "app_state"

    id = db.Column(db.Integer, primary_key=True)
    state_key = db.Column(db.String(120), nullable=False, unique=True)
    state_value = db.Column(db.Text, nullable=False, default="")


class AnalysisType(db.Model):
    __tablename__ = "analysis_types"

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False, unique=True)
    created_at = db.Column(db.DateTime, server_default=func.now(), nullable=False)
    updated_at = db.Column(db.DateTime, server_default=func.now(), onupdate=func.now(), nullable=False)
    charts = db.relationship(
        "AnalysisChart",
        backref="analysis_type",
        cascade="all, delete-orphan",
        order_by="AnalysisChart.position",
        lazy=True,
    )


class AnalysisChart(db.Model):
    __tablename__ = "analysis_charts"

    id = db.Column(db.Integer, primary_key=True)
    analysis_type_id = db.Column(db.Integer, db.ForeignKey("analysis_types.id"), nullable=False)
    chart_type = db.Column(db.String(40), nullable=False, default="bar")
    source_kind = db.Column(db.String(40), nullable=False, default="none")
    x_field = db.Column(db.String(120), nullable=False, default="")
    y_field = db.Column(db.String(120), nullable=False, default="")
    group_field = db.Column(db.String(120), nullable=False, default="")
    agg_func = db.Column(db.String(40), nullable=False, default="count")
    color = db.Column(db.String(40), nullable=False, default="#b7791f")
    legend = db.Column(db.String(120), nullable=False, default="")
    labels = db.Column(db.String(120), nullable=False, default="")
    comment_title = db.Column(db.String(255), nullable=False, default="")
    comment_text = db.Column(db.Text, nullable=False, default="")
    is_hidden = db.Column(db.Boolean, nullable=False, default=False)
    position = db.Column(db.Integer, nullable=False, default=0)


class AnalysisUserState(db.Model):
    __tablename__ = "analysis_user_state"

    id = db.Column(db.Integer, primary_key=True)
    selected_analysis_type_id = db.Column(db.Integer, db.ForeignKey("analysis_types.id"), nullable=True)
    left_panel_width = db.Column(db.Integer, nullable=False, default=260)
    visual_source_kind = db.Column(db.String(40), nullable=False, default="none")
    table_search = db.Column(db.Text, nullable=False, default="")
    table_sort_column = db.Column(db.String(120), nullable=False, default="")
    table_sort_direction = db.Column(db.String(8), nullable=False, default="asc")
    draft_state_json = db.Column(db.Text, nullable=False, default="")


class AnalysisDataSource(db.Model):
    __tablename__ = "analysis_data_sources"

    id = db.Column(db.Integer, primary_key=True)
    source_kind = db.Column(db.String(40), nullable=False, default="none")
    source_mode = db.Column(db.String(40), nullable=False, default="none")
    source_label = db.Column(db.String(255), nullable=False, default="")
    source_file_name = db.Column(db.String(255), nullable=False, default="")
    attached_at = db.Column(db.DateTime, server_default=func.now(), nullable=False)
    is_active = db.Column(db.Boolean, nullable=False, default=False)
    is_available = db.Column(db.Boolean, nullable=False, default=True)
    row_count = db.Column(db.Integer, nullable=False, default=0)
    column_count = db.Column(db.Integer, nullable=False, default=0)
    payload_json = db.Column(db.Text, nullable=False, default="{}")


class EventLog(db.Model):
    __tablename__ = "event_log"

    id = db.Column(db.Integer, primary_key=True)
    event_type = db.Column(db.String(80), nullable=False)
    event_description = db.Column(db.Text, nullable=False, default="")
    created_at = db.Column(db.DateTime, server_default=func.now(), nullable=False)


def serialize_template(template):
    return {
        "id": template.id,
        "name": template.name,
        "created_at": template.created_at.isoformat() if template.created_at else None,
        "columns": [
            {
                "id": column.id,
                "name": column.column_name,
                "type": column.column_type,
                "position": column.position,
            }
            for column in template.columns
        ],
    }


def serialize_analysis_chart(chart):
    return {
        "id": chart.id,
        "chart_type": chart.chart_type,
        "source_kind": chart.source_kind,
        "x_field": chart.x_field,
        "y_field": chart.y_field,
        "group_field": chart.group_field,
        "agg_func": chart.agg_func,
        "color": chart.color,
        "legend": chart.legend,
        "labels": chart.labels,
        "comment_title": chart.comment_title,
        "comment_text": chart.comment_text,
        "is_hidden": chart.is_hidden,
        "position": chart.position,
    }


def serialize_analysis_type(item):
    return {
        "id": item.id,
        "name": item.name,
        "created_at": item.created_at.isoformat() if item.created_at else None,
        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
        "charts": [serialize_analysis_chart(chart) for chart in item.charts],
    }


def get_active_analysis_source():
    return AnalysisDataSource.query.filter_by(is_active=True).order_by(AnalysisDataSource.id.desc()).first()


def deactivate_active_analysis_sources():
    AnalysisDataSource.query.filter_by(is_active=True).update({"is_active": False}, synchronize_session=False)


def load_analysis_dataset_from_source(source):
    if source is None:
        return deepcopy(EMPTY_ANALYSIS_DATASET)

    try:
        payload = json.loads(source.payload_json or "{}")
    except json.JSONDecodeError:
        payload = {}

    dataset = deepcopy(EMPTY_ANALYSIS_DATASET)
    dataset.update(payload if isinstance(payload, dict) else {})
    dataset["source_kind"] = source.source_kind or dataset["source_kind"]
    dataset["source_mode"] = source.source_mode or dataset["source_mode"]
    dataset["source_label"] = source.source_label or dataset["source_label"]
    dataset["source_file_name"] = source.source_file_name or dataset["source_file_name"]
    dataset["attached_at"] = source.attached_at.isoformat() if source.attached_at else dataset["attached_at"]
    dataset["source_status"] = "ready" if source.is_available else "unavailable"
    dataset["row_count"] = source.row_count
    dataset["column_count"] = source.column_count
    dataset.setdefault("warnings", [])
    return dataset


def persist_analysis_source(dataset):
    deactivate_active_analysis_sources()
    source = AnalysisDataSource(
        source_kind=dataset["source_kind"],
        source_mode=dataset.get("source_mode") or "none",
        source_label=dataset.get("source_label") or "",
        source_file_name=dataset.get("source_file_name") or "",
        is_active=True,
        is_available=True,
        row_count=dataset.get("row_count") or len(dataset.get("rows") or []),
        column_count=dataset.get("column_count") or len(dataset.get("columns") or []),
        payload_json=json.dumps(dataset, ensure_ascii=False),
    )
    db.session.add(source)
    CURRENT_ANALYSIS_DATASET.clear()
    CURRENT_ANALYSIS_DATASET.update(dataset)
    return source


def get_analysis_user_state():
    state = AnalysisUserState.query.order_by(AnalysisUserState.id).first()
    if state is None:
        return dict(DEFAULT_ANALYSIS_USER_STATE)

    draft_state = None
    if state.draft_state_json:
        try:
            raw_draft_state = json.loads(state.draft_state_json)
        except json.JSONDecodeError:
            raw_draft_state = None

        if isinstance(raw_draft_state, dict):
            analysis_type_id = raw_draft_state.get("analysis_type_id")
            charts, _ = normalize_analysis_charts(raw_draft_state.get("charts"))
            draft_state = {
                "analysis_type_id": analysis_type_id,
                "name": str(raw_draft_state.get("name") or "").strip(),
                "charts": charts,
            }

    return {
        "selected_analysis_type_id": state.selected_analysis_type_id,
        "left_panel_width": state.left_panel_width,
        "visual_source_kind": state.visual_source_kind or "none",
        "table_search": state.table_search or "",
        "table_sort_column": state.table_sort_column or "",
        "table_sort_direction": state.table_sort_direction or "asc",
        "draft_state": draft_state,
    }


def upsert_analysis_user_state(payload):
    state = AnalysisUserState.query.order_by(AnalysisUserState.id).first()
    if state is None:
        state = AnalysisUserState(**DEFAULT_ANALYSIS_USER_STATE)
        db.session.add(state)
        db.session.flush()

    if "selected_analysis_type_id" in payload:
        state.selected_analysis_type_id = payload["selected_analysis_type_id"]
    if "left_panel_width" in payload:
        state.left_panel_width = payload["left_panel_width"]
    if "visual_source_kind" in payload:
        state.visual_source_kind = payload["visual_source_kind"]
    if "table_search" in payload:
        state.table_search = str(payload["table_search"] or "")
    if "table_sort_column" in payload:
        state.table_sort_column = str(payload["table_sort_column"] or "")
    if "table_sort_direction" in payload:
        state.table_sort_direction = payload["table_sort_direction"]
    if "draft_state" in payload:
        draft_state = payload["draft_state"]
        state.draft_state_json = json.dumps(draft_state, ensure_ascii=False) if draft_state else ""

    return state


def ensure_analysis_user_state_schema():
    required_columns = {
        "table_search": "ALTER TABLE analysis_user_state ADD COLUMN table_search TEXT NOT NULL DEFAULT ''",
        "table_sort_column": "ALTER TABLE analysis_user_state ADD COLUMN table_sort_column VARCHAR(120) NOT NULL DEFAULT ''",
        "table_sort_direction": "ALTER TABLE analysis_user_state ADD COLUMN table_sort_direction VARCHAR(8) NOT NULL DEFAULT 'asc'",
        "draft_state_json": "ALTER TABLE analysis_user_state ADD COLUMN draft_state_json TEXT NOT NULL DEFAULT ''",
    }

    with db.engine.begin() as connection:
        table_names = {row[0] for row in connection.exec_driver_sql("SELECT name FROM sqlite_master WHERE type='table'")}
        if "analysis_user_state" not in table_names:
            return

        existing_columns = {row[1] for row in connection.exec_driver_sql("PRAGMA table_info(analysis_user_state)")}
        for column_name, statement in required_columns.items():
            if column_name not in existing_columns:
                connection.exec_driver_sql(statement)


def get_settings_map():
    settings = {key: "" for key in SUPPORTED_SETTING_KEYS}
    for item in AppSetting.query.order_by(AppSetting.key).all():
        if item.key in settings:
            settings[item.key] = item.value
    return settings


def get_state_map():
    state = dict(DEFAULT_STATE)
    for item in AppState.query.all():
        try:
            state[item.state_key] = json.loads(item.state_value)
        except json.JSONDecodeError:
            state[item.state_key] = item.state_value
    return state


def upsert_setting(key, value):
    item = AppSetting.query.filter_by(key=key).first()
    if item is None:
        item = AppSetting(key=key, value=value)
        db.session.add(item)
    else:
        item.value = value


def upsert_state(key, value):
    item = AppState.query.filter_by(state_key=key).first()
    serialized = json.dumps(value, ensure_ascii=False)
    if item is None:
        item = AppState(state_key=key, state_value=serialized)
        db.session.add(item)
    else:
        item.state_value = serialized


def log_event(event_type, description):
    db.session.add(EventLog(event_type=event_type, event_description=description))


def infer_event_level(event_type):
    if event_type.endswith("_failed") or "error" in event_type:
        return "error"
    if event_type.endswith("_partial") or "warning" in event_type:
        return "warning"
    return "info"


def get_column_type_label(column_type):
    return COLUMN_TYPE_LABELS.get(column_type, column_type)


def serialize_log_entry(item):
    return {
        "id": item.id,
        "event_level": infer_event_level(item.event_type),
        "event_type": item.event_type,
        "event_description": item.event_description,
        "created_at": item.created_at.isoformat() if item.created_at else None,
    }


def build_logs_payload(limit=80, current_limit=20, history_limit=30):
    entries = EventLog.query.order_by(EventLog.created_at.desc(), EventLog.id.desc()).limit(limit).all()
    current_session = []
    history = []

    for entry in entries:
        target = current_session if entry.created_at and entry.created_at >= APP_SESSION_STARTED_AT else history
        if target is current_session and len(current_session) >= current_limit:
            continue
        if target is history and len(history) >= history_limit:
            continue
        target.append(serialize_log_entry(entry))

        if len(current_session) >= current_limit and len(history) >= history_limit:
            break

    return {
        "session_started_at": APP_SESSION_STARTED_AT.isoformat(),
        "current_session": current_session,
        "history": history,
    }


def read_instructions_file(file_path):
    path = Path(file_path)
    if not path.exists() or not path.is_file():
        raise ValueError("Файл инструкций недоступен. Проверьте путь в настройках.")

    try:
        content = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("Файл инструкций должен быть текстовым в UTF-8.") from exc

    if "\x00" in content:
        raise ValueError("Файл инструкций должен быть plain text.")

    return {
        "file_name": path.name,
        "file_path": str(path),
        "content": content,
    }


def log_dataset_quality(dataset, source_event):
    warning_count = len(dataset["warnings"])
    error_count = len(dataset["errors"])

    if warning_count:
        log_event(
            "system_warning",
            f"{source_event}: получено предупреждений {warning_count}.",
        )

    if error_count:
        log_event(
            "typing_error",
            f"{source_event}: найдено ошибок типизации {error_count}.",
        )


def serialize_dataset():
    payload = deepcopy(CURRENT_DATASET)
    payload.pop("next_row_id", None)
    return payload


def serialize_analysis_dataset():
    if CURRENT_ANALYSIS_DATASET["source_kind"] != "none" or CURRENT_ANALYSIS_DATASET["columns"]:
        return deepcopy(CURRENT_ANALYSIS_DATASET)
    return load_analysis_dataset_from_source(get_active_analysis_source())


def build_analysis_dataset_from_current_facts_export_layer():
    if not CURRENT_DATASET["columns"] or not CURRENT_DATASET["rows"]:
        raise ValueError("Во вкладке 'Загрузка фактов' нет данных для анализа.")

    ordered_columns = sorted(CURRENT_DATASET["columns"], key=lambda item: item.get("position", 0))
    column_names = [column["name"] for column in ordered_columns]
    rows = []
    for index, row in enumerate(CURRENT_DATASET["rows"], start=1):
        rows.append(
            {
                "row_number": index,
                "source_row_number": row.get("source_row_number"),
                "values": {column_name: row.get("values", {}).get(column_name, "") for column_name in column_names},
            }
        )

    return {
        "source_kind": "facts",
        "source_mode": "export_layer",
        "source_label": CURRENT_DATASET.get("file_name") or "Текущая загрузка фактов",
        "source_file_name": CURRENT_DATASET.get("file_name") or "",
        "attached_at": datetime.utcnow().replace(microsecond=0).isoformat(),
        "source_status": "ready",
        "columns": deepcopy(ordered_columns),
        "rows": rows,
        "row_count": len(rows),
        "column_count": len(ordered_columns),
        "warnings": [],
    }


def build_analysis_dataset_from_uploaded_file(file_storage):
    dataset = build_dataset(file_storage, None)
    return {
        "source_kind": "file",
        "source_mode": "uploaded_file",
        "source_label": dataset.get("file_name") or "Файл анализа",
        "source_file_name": dataset.get("file_name") or "",
        "attached_at": datetime.utcnow().replace(microsecond=0).isoformat(),
        "source_status": "ready",
        "columns": deepcopy(dataset.get("columns") or []),
        "rows": [
            {
                "row_number": index,
                "source_row_number": row.get("source_row_number"),
                "values": deepcopy(row.get("values") or {}),
            }
            for index, row in enumerate(dataset.get("rows") or [], start=1)
        ],
        "row_count": dataset.get("row_count") or len(dataset.get("rows") or []),
        "column_count": len(dataset.get("columns") or []),
        "warnings": deepcopy(dataset.get("warnings") or []),
    }


def build_default_analysis_chart(position):
    return {
        "chart_type": "bar",
        "source_kind": "none",
        "x_field": "",
        "y_field": "",
        "group_field": "",
        "agg_func": "count",
        "color": "#b7791f",
        "legend": "",
        "labels": "",
        "comment_title": "",
        "comment_text": "",
        "is_hidden": False,
        "position": position,
    }


def normalize_analysis_charts(raw_charts):
    charts = []
    errors = []

    if raw_charts is None:
        return charts, errors

    if not isinstance(raw_charts, list):
        return [], ["Карточки графиков переданы в некорректном формате."]

    if len(raw_charts) > 4:
        return [], ["Во вкладке 'Анализ' допускается не более 4 графиков."]

    for index, item in enumerate(raw_charts):
        if not isinstance(item, dict):
            errors.append(f"График #{index + 1}: некорректный формат карточки.")
            continue

        chart_type = (item.get("chart_type") or "bar").strip().lower()
        source_kind = (item.get("source_kind") or "none").strip().lower()
        agg_func = (item.get("agg_func") or "count").strip().lower() or "count"
        if chart_type not in ALLOWED_ANALYSIS_CHART_TYPES:
            errors.append(f"График #{index + 1}: тип '{chart_type}' не поддерживается.")
            continue
        if source_kind not in ALLOWED_ANALYSIS_SOURCES:
            errors.append(f"График #{index + 1}: источник '{source_kind}' не поддерживается.")
            continue
        if agg_func not in ALLOWED_ANALYSIS_AGG_FUNCS:
            errors.append(f"График #{index + 1}: агрегация '{agg_func}' не поддерживается.")
            continue

        charts.append(
            {
                "chart_type": chart_type,
                "source_kind": source_kind,
                "x_field": (item.get("x_field") or "").strip(),
                "y_field": (item.get("y_field") or "").strip(),
                "group_field": (item.get("group_field") or "").strip(),
                "agg_func": agg_func,
                "color": (item.get("color") or "#b7791f").strip() or "#b7791f",
                "legend": (item.get("legend") or "").strip(),
                "labels": (item.get("labels") or "").strip(),
                "comment_title": (item.get("comment_title") or "").strip()[:255],
                "comment_text": str(item.get("comment_text") or "").strip(),
                "is_hidden": bool(item.get("is_hidden")),
                "position": len(charts),
            }
        )

    return charts, errors


def validate_analysis_type_payload(payload, type_id=None):
    errors = []
    name = (payload.get("name") or "").strip()
    charts, chart_errors = normalize_analysis_charts(payload.get("charts"))

    if not name:
        errors.append("Укажите имя типа анализа.")

    duplicate_query = AnalysisType.query.filter(func.lower(AnalysisType.name) == name.lower())
    if type_id is not None:
        duplicate_query = duplicate_query.filter(AnalysisType.id != type_id)
    if name and duplicate_query.first():
        errors.append("Тип анализа с таким именем уже существует.")

    errors.extend(chart_errors)
    return name, charts, errors


def replace_analysis_type_charts(analysis_type, charts):
    AnalysisChart.query.filter_by(analysis_type_id=analysis_type.id).delete()
    for item in charts:
        db.session.add(
            AnalysisChart(
                analysis_type_id=analysis_type.id,
                chart_type=item["chart_type"],
                source_kind=item["source_kind"],
                x_field=item["x_field"],
                y_field=item["y_field"],
                group_field=item["group_field"],
                agg_func=item["agg_func"],
                color=item["color"],
                legend=item["legend"],
                labels=item["labels"],
                comment_title=item["comment_title"],
                comment_text=item["comment_text"],
                is_hidden=item["is_hidden"],
                position=item["position"],
            )
        )


def ensure_analysis_chart_schema():
    required_columns = {
        "comment_title": "ALTER TABLE analysis_charts ADD COLUMN comment_title VARCHAR(255) NOT NULL DEFAULT ''",
        "comment_text": "ALTER TABLE analysis_charts ADD COLUMN comment_text TEXT NOT NULL DEFAULT ''",
    }

    with db.engine.begin() as connection:
        table_names = {row[0] for row in connection.exec_driver_sql("SELECT name FROM sqlite_master WHERE type='table'")}
        if "analysis_charts" not in table_names:
            return

        existing_columns = {
            row[1] for row in connection.exec_driver_sql("PRAGMA table_info(analysis_charts)")
        }
        for column_name, statement in required_columns.items():
            if column_name not in existing_columns:
                connection.exec_driver_sql(statement)


def build_analysis_preview_state(state, message, **extra):
    payload = {
        "state": state,
        "message": message,
        "chart_type": extra.pop("chart_type", None),
        "warnings": extra.pop("warnings", []),
        "summary": extra.pop("summary", {}),
    }
    payload.update(extra)
    return payload


def get_analysis_column_map(dataset):
    return {str(column.get("name") or ""): column for column in dataset.get("columns") or []}


def get_analysis_column(dataset, field_name):
    return get_analysis_column_map(dataset).get(field_name)


def normalize_analysis_numeric_value(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    normalized = text.replace(" ", "").replace(",", ".")
    try:
        return float(normalized)
    except ValueError:
        return None


def is_analysis_numeric_column(column):
    return (column or {}).get("type") in {"number", "money"}


def can_analysis_column_be_used_as_numeric(dataset, field_name, column):
    if is_analysis_numeric_column(column):
        return True

    saw_non_empty_value = False
    for row in dataset.get("rows") or []:
        values = row.get("values") or {}
        raw_value = values.get(field_name, "")
        if str(raw_value or "").strip():
            saw_non_empty_value = True
        if normalize_analysis_numeric_value(raw_value) is not None:
            return True

    return not saw_non_empty_value


def pick_analysis_series_color(base_color, index):
    palette = [
        base_color or "#b7791f",
        "#0f766e",
        "#c2410c",
        "#2563eb",
        "#7c3aed",
        "#be123c",
    ]
    return palette[index % len(palette)]


def build_analysis_chart_preview(chart, dataset):
    normalized_charts, errors = normalize_analysis_charts([chart])
    if errors:
        return build_analysis_preview_state("error", " ".join(errors), chart_type=(chart or {}).get("chart_type") or "bar")

    if not normalized_charts:
        return build_analysis_preview_state("error", "Некорректная конфигурация графика.", chart_type=(chart or {}).get("chart_type") or "bar")

    chart = normalized_charts[0]
    chart_type = chart["chart_type"]
    source_kind = chart["source_kind"]
    active_source_kind = dataset.get("source_kind") or "none"
    columns = dataset.get("columns") or []
    rows = dataset.get("rows") or []

    if chart.get("is_hidden"):
        return build_analysis_preview_state("hidden", "График скрыт пользователем.", chart_type=chart_type)

    if not columns or not rows:
        return build_analysis_preview_state("empty", "Источник данных не подключен или не содержит строк.", chart_type=chart_type)

    if source_kind != "none" and active_source_kind != "none" and source_kind != active_source_kind:
        return build_analysis_preview_state(
            "incompatible",
            "Источник карточки не совпадает с активным источником данных Анализа.",
            chart_type=chart_type,
        )

    x_field = chart.get("x_field") or ""
    y_field = chart.get("y_field") or ""
    group_field = chart.get("group_field") or ""
    agg_func = chart.get("agg_func") or "count"

    if not x_field:
        return build_analysis_preview_state("needs-config", "Выберите поле X для построения графика.", chart_type=chart_type)

    x_column = get_analysis_column(dataset, x_field)
    if x_column is None:
        return build_analysis_preview_state("incompatible", f"Поле X '{x_field}' отсутствует в текущем источнике.", chart_type=chart_type)

    y_column = get_analysis_column(dataset, y_field) if y_field else None
    if agg_func != "count" and not y_field:
        return build_analysis_preview_state("needs-config", "Для выбранной агрегации нужно указать поле Y.", chart_type=chart_type)
    if y_field and y_column is None:
        return build_analysis_preview_state("incompatible", f"Поле Y '{y_field}' отсутствует в текущем источнике.", chart_type=chart_type)
    if agg_func in {"sum", "avg", "min", "max"} and not can_analysis_column_be_used_as_numeric(dataset, y_field, y_column):
        return build_analysis_preview_state("error", f"Агрегация '{agg_func}' требует числовое поле Y.", chart_type=chart_type)

    group_column = get_analysis_column(dataset, group_field) if group_field else None
    if group_field and group_column is None:
        return build_analysis_preview_state("incompatible", f"Поле Group '{group_field}' отсутствует в текущем источнике.", chart_type=chart_type)

    categories = []
    category_index = {}
    series_order = []
    series_buckets = {}
    skipped_rows = 0

    def ensure_category(value):
        key = str(value or "Пусто")
        if key not in category_index:
            category_index[key] = len(categories)
            categories.append(key)
        return key

    def ensure_series(value):
        key = str(value or "Без группы") if group_field else (chart.get("legend") or "Значение")
        if key not in series_buckets:
            series_buckets[key] = {}
            series_order.append(key)
        return key

    for row in rows:
        values = row.get("values") or {}
        category_key = ensure_category(values.get(x_field, ""))
        series_key = ensure_series(values.get(group_field, "") if group_field else None)
        bucket = series_buckets[series_key].setdefault(category_key, {"sum": 0.0, "count": 0, "min": None, "max": None})

        if agg_func == "count":
            bucket["count"] += 1
            continue

        numeric_value = normalize_analysis_numeric_value(values.get(y_field, ""))
        if numeric_value is None:
            skipped_rows += 1
            continue

        bucket["count"] += 1
        bucket["sum"] += numeric_value
        bucket["min"] = numeric_value if bucket["min"] is None else min(bucket["min"], numeric_value)
        bucket["max"] = numeric_value if bucket["max"] is None else max(bucket["max"], numeric_value)

    series = []
    for index, series_name in enumerate(series_order):
        bucket_map = series_buckets.get(series_name) or {}
        values = []
        for category in categories:
            bucket = bucket_map.get(category)
            if not bucket:
                values.append(0)
                continue
            if agg_func == "count":
                values.append(bucket["count"])
            elif agg_func == "sum":
                values.append(round(bucket["sum"], 4))
            elif agg_func == "avg":
                values.append(round(bucket["sum"] / bucket["count"], 4) if bucket["count"] else 0)
            elif agg_func == "min":
                values.append(round(bucket["min"], 4) if bucket["min"] is not None else 0)
            elif agg_func == "max":
                values.append(round(bucket["max"], 4) if bucket["max"] is not None else 0)
            else:
                values.append(0)

        series.append(
            {
                "name": series_name,
                "color": pick_analysis_series_color(chart.get("color"), index),
                "values": values,
            }
        )

    non_zero_points = sum(1 for item in series for value in item["values"] if value not in (0, 0.0, None))
    if not series or not categories or non_zero_points == 0:
        return build_analysis_preview_state(
            "no-data",
            "По текущей конфигурации нет данных для построения графика.",
            chart_type=chart_type,
            warnings=[f"Пропущено строк из-за неподходящих значений: {skipped_rows}."] if skipped_rows else [],
            summary={"aggregation": agg_func, "skipped_rows": skipped_rows, "points": 0},
        )

    table_rows = []
    for category_position, category in enumerate(categories):
        row_payload = {"category": category}
        for item in series:
            row_payload[item["name"]] = item["values"][category_position]
        table_rows.append(row_payload)

    warnings = []
    if skipped_rows:
        warnings.append(f"Часть строк пропущена из-за неподходящих значений: {skipped_rows}.")

    return build_analysis_preview_state(
        "ready",
        "График построен.",
        chart_type=chart_type,
        categories=categories,
        series=series,
        table_rows=table_rows,
        x_label=x_field,
        y_label=y_field or agg_func,
        group_label=group_field,
        legend=chart.get("legend") or "",
        labels=chart.get("labels") or "",
        warnings=warnings,
        summary={
            "aggregation": agg_func,
            "skipped_rows": skipped_rows,
            "points": non_zero_points,
            "source_kind": active_source_kind,
        },
    )


def build_dataset_errors(rows, columns):
    errors = []
    for row in rows:
        for column in columns:
            value = row["values"].get(column["name"], "")
            reason = validate_value(value, column["type"])
            if reason:
                errors.append(
                    {
                        "row_number": row["row_number"],
                        "source_row_number": row["source_row_number"],
                        "column_name": column["name"],
                        "column_type": column["type"],
                        "value": value,
                        "reason": reason,
                    }
                )
    return errors


def rebuild_row_numbers(rows):
    for index, row in enumerate(rows, start=1):
        row["row_number"] = index


def split_parse_fragments(value, delimiter):
    return [fragment.strip() for fragment in str(value or "").split(delimiter) if fragment.strip()]


def build_export_path(output_dir):
    base_name = datetime.now().strftime("data_%d%m%y_%H%M")
    candidate = output_dir / f"{base_name}.xlsx"
    suffix = 2

    while candidate.exists():
        candidate = output_dir / f"{base_name}_{suffix}.xlsx"
        suffix += 1

    return candidate


def sanitize_export_stem(value, fallback):
    normalized = re.sub(r"[^\w.-]+", "_", str(value or "").strip(), flags=re.UNICODE).strip("._")
    return normalized or fallback


def build_named_export_path(output_dir, stem, extension):
    base_name = sanitize_export_stem(stem, f"export_{datetime.now().strftime('%d%m%y_%H%M')}")
    candidate = output_dir / f"{base_name}.{extension}"
    suffix = 2

    while candidate.exists():
        candidate = output_dir / f"{base_name}_{suffix}.{extension}"
        suffix += 1

    return candidate


def get_output_dir_error():
    settings = get_settings_map()
    output_folder = (settings.get("output_folder") or "").strip()
    if not output_folder:
        return None, "Не задана папка результата в настройках."

    output_dir = Path(output_folder)
    if not output_dir.exists() or not output_dir.is_dir():
        return None, "Папка результата недоступна. Проверьте настройки приложения."

    return output_dir, None


def sanitize_worksheet_title(value):
    title = re.sub(r"[\\/*?:\[\]]+", "_", str(value or "data")).strip()
    return (title or "data")[:31]


def save_rows_to_xlsx(target_path, columns, raw_rows, sheet_title="data"):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sanitize_worksheet_title(sheet_title)
    worksheet.append(columns)

    for row in raw_rows:
        worksheet.append([str((row or {}).get(column_name, "")) for column_name in columns])

    workbook.save(target_path)


def normalize_header_name(raw_name, index):
    normalized = (raw_name or "").strip()
    if normalized:
        return normalized
    return f"Столбец_{index + 1}"


def uniquify_headers(headers):
    seen = {}
    result = []
    warnings = []

    for index, raw_name in enumerate(headers):
        base_name = normalize_header_name(raw_name, index)
        count = seen.get(base_name.lower(), 0) + 1
        seen[base_name.lower()] = count
        if count == 1:
            unique_name = base_name
        else:
            unique_name = f"{base_name}_{count}"
            warnings.append(
                f"Дублирующийся столбец '{base_name}' переименован в '{unique_name}'."
            )
        if not (raw_name or "").strip():
            warnings.append(f"Пустой заголовок столбца заменён на '{unique_name}'.")
        result.append(unique_name)

    return result, warnings


def parse_decimal(value):
    normalized = str(value).strip()
    if not normalized:
        return None

    if not re.fullmatch(r"[+-]?(?:\d+|\d{1,3}(?: \d{3})+)(?:\.\d+)?", normalized):
        raise ValueError("invalid decimal")

    return float(normalized.replace(" ", ""))


def normalize_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def load_csv_rows(file_storage):
    try:
        decoded = file_storage.read().decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV-файл должен быть в кодировке UTF-8 with BOM.") from exc

    return [list(row) for row in csv.reader(io.StringIO(decoded), delimiter=";")]


def load_xlsx_rows(file_storage):
    try:
        workbook = load_workbook(filename=io.BytesIO(file_storage.read()), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Не удалось прочитать XLSX-файл.") from exc

    worksheet = workbook.active
    try:
        return [
            [normalize_cell_value(cell) for cell in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
    finally:
        workbook.close()


def parse_boolean(value):
    normalized = str(value).strip().lower()
    if normalized in {"true", "yes", "да", "1", "истина"}:
        return True
    if normalized in {"false", "no", "нет", "0", "ложь"}:
        return False
    raise ValueError("invalid boolean")


def validate_value(value, column_type):
    raw_value = (value or "").strip()
    if column_type in {"text", "choice"}:
        return None

    try:
        if column_type in {"number", "money"}:
            if not raw_value:
                return "пустое значение трактуется как 0"
            parse_decimal(raw_value)
            return None
        if column_type == "boolean":
            if not raw_value:
                return "пустое значение трактуется как Нет"
            parse_boolean(raw_value)
            return None
        if column_type == "date":
            if not raw_value:
                return "пустое значение трактуется как 01.01.1900"
            for pattern in ("%Y-%m-%d", "%d.%m.%Y"):
                try:
                    datetime.strptime(raw_value, pattern)
                    return None
                except ValueError:
                    continue
            return "ожидалась дата"
        if column_type == "datetime":
            if not raw_value:
                return "ожидались дата и время"
            for pattern in (
                "%d.%m.%Y %H:%M",
                "%Y-%m-%d %H:%M",
                "%Y-%m-%dT%H:%M:%S",
            ):
                try:
                    datetime.strptime(raw_value, pattern)
                    return None
                except ValueError:
                    continue
            return "ожидались дата и время"
    except ValueError:
        if column_type in {"number", "money"}:
            return "ожидалось числовое значение"
        if column_type == "boolean":
            return "ожидалось логическое значение"

    return None


def build_dataset(file_storage, template):
    if file_storage is None:
        raise ValueError("Файл не выбран.")

    file_name = Path(file_storage.filename or "").name
    if not file_name:
        raise ValueError("Файл не выбран.")
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        source_rows = load_csv_rows(file_storage)
    elif suffix == ".xlsx":
        source_rows = load_xlsx_rows(file_storage)
    else:
        raise ValueError("Поддерживаются только CSV- и XLSX-файлы.")

    header = None
    header_row_number = None
    warnings = []
    rows = []

    for row_number, row in enumerate(source_rows, start=1):
        if header is None and not any((cell or "").strip() for cell in row):
            continue
        if header is None:
            header_row_number = row_number
            header, header_warnings = uniquify_headers(row)
            warnings.extend(header_warnings)
            continue

        values = list(row)
        if len(values) < len(header):
            values.extend([""] * (len(header) - len(values)))
        elif len(values) > len(header):
            warnings.append(
                f"Строка {row_number}: лишние значения после столбца '{header[-1]}' были отброшены."
            )
            values = values[: len(header)]

        rows.append(
            {
                "row_id": len(rows) + 1,
                "row_number": len(rows) + 1,
                "source_row_number": row_number,
                "values": {column_name: values[index] for index, column_name in enumerate(header)},
            }
        )

    if header is None:
        raise ValueError("В файле не найден заголовок.")

    template_columns = template.columns if template else []
    template_types = {column.column_name.lower(): column.column_type for column in template_columns}
    expected_columns = [column.column_name for column in template_columns]
    missing_columns = [name for name in expected_columns if name.lower() not in {item.lower() for item in header}]
    partial = bool(missing_columns)
    if missing_columns:
        warnings.append(
            "Частичная загрузка: отсутствуют ожидаемые столбцы шаблона: " + ", ".join(missing_columns)
        )

    known_template_columns = {column.column_name.lower() for column in template_columns}
    for column_name in header:
        if template and column_name.lower() not in known_template_columns:
            warnings.append(
                f"Новый столбец '{column_name}' не найден в шаблоне и получил тип '{get_column_type_label('text')}' по умолчанию."
            )

    columns = []
    for position, column_name in enumerate(header):
        column_type = template_types.get(column_name.lower(), "text")
        columns.append(
            {
                "name": column_name,
                "type": column_type,
                "position": position,
                "from_template": column_name.lower() in template_types,
            }
        )

    errors = build_dataset_errors(rows, columns)

    return {
        "file_name": file_name,
        "header_row_number": header_row_number,
        "columns": columns,
        "rows": rows,
        "warnings": warnings,
        "errors": errors,
        "partial": partial,
        "missing_columns": missing_columns,
        "row_count": len(rows),
        "template": {"id": template.id, "name": template.name} if template else None,
        "next_row_id": len(rows) + 1,
        "can_undo_parse": False,
    }


def validate_settings(payload):
    errors = {}
    folder_fields = {
        "output_folder": "Папка результата",
    }
    file_fields = {
        "instructions_file": "Файл инструкций",
    }

    for key, label in folder_fields.items():
        value = (payload.get(key) or "").strip()
        if value and (not Path(value).exists() or not Path(value).is_dir()):
            errors[key] = f"{label} должна существовать и быть папкой."

    for key, label in file_fields.items():
        value = (payload.get(key) or "").strip()
        if value and (not Path(value).exists() or not Path(value).is_file()):
            errors[key] = f"{label} должен существовать и быть файлом."

    return errors


def normalize_columns(raw_columns):
    columns = []
    errors = []
    seen_names = set()

    for index, item in enumerate(raw_columns or []):
        name = (item.get("name") or "").strip()
        column_type = (item.get("type") or "text").strip().lower()
        if not name:
            errors.append(f"Столбец #{index + 1}: укажите название.")
            continue
        if name.lower() in seen_names:
            errors.append(f"Столбец '{name}' указан несколько раз.")
            continue
        if column_type not in ALLOWED_COLUMN_TYPES:
            errors.append(f"Столбец '{name}': тип '{get_column_type_label(column_type)}' не поддерживается.")
            continue
        seen_names.add(name.lower())
        columns.append({"name": name, "type": column_type, "position": len(columns)})

    return columns, errors


@app.route("/")
def home():
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/styles.css")
def styles():
    return send_from_directory(BASE_DIR, "styles.css")


@app.route("/app.js")
def script():
    return send_from_directory(BASE_DIR, "app.js")


@app.route("/analysis.js")
def analysis_script():
    return send_from_directory(BASE_DIR, "analysis.js")


@app.get("/api/bootstrap")
def bootstrap():
    return jsonify(
        {
            "settings": get_settings_map(),
            "templates": [serialize_template(item) for item in Template.query.order_by(Template.name).all()],
            "state": get_state_map(),
            "dataset": serialize_dataset(),
            "logs": build_logs_payload(),
        }
    )


@app.get("/api/settings")
def get_settings():
    return jsonify(get_settings_map())


@app.put("/api/settings")
def save_settings():
    payload = request.get_json(silent=True) or {}
    errors = validate_settings(payload)
    if errors:
        log_event("settings_failed", "Ошибка валидации настроек приложения.")
        db.session.commit()
        return jsonify({"ok": False, "errors": errors}), 400

    for key in SUPPORTED_SETTING_KEYS:
        upsert_setting(key, (payload.get(key) or "").strip())

    log_event("settings_saved", "Настройки приложения обновлены.")
    db.session.commit()
    return jsonify({"ok": True, "settings": get_settings_map()})


@app.get("/api/templates")
def list_templates():
    templates = Template.query.order_by(Template.name).all()
    return jsonify([serialize_template(item) for item in templates])


@app.post("/api/templates")
def create_template():
    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or "").strip()
    columns, column_errors = normalize_columns(payload.get("columns") or [])

    errors = []
    if not name:
        errors.append("Укажите имя шаблона.")
    if Template.query.filter(func.lower(Template.name) == name.lower()).first():
        errors.append("Шаблон с таким именем уже существует.")
    errors.extend(column_errors)
    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    template = Template(name=name)
    db.session.add(template)
    db.session.flush()

    for column in columns:
        db.session.add(
            TemplateColumn(
                template_id=template.id,
                column_name=column["name"],
                column_type=column["type"],
                position=column["position"],
            )
        )

    log_event("template_created", f"Создан шаблон '{name}'.")
    db.session.commit()
    return jsonify({"ok": True, "template": serialize_template(template)}), 201


@app.put("/api/templates/<int:template_id>")
def update_template(template_id):
    template = Template.query.get_or_404(template_id)
    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or "").strip()
    columns, column_errors = normalize_columns(payload.get("columns") or [])

    errors = []
    if not name:
        errors.append("Укажите имя шаблона.")

    duplicate = Template.query.filter(func.lower(Template.name) == name.lower(), Template.id != template_id).first()
    if duplicate:
        errors.append("Шаблон с таким именем уже существует.")

    errors.extend(column_errors)
    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    template.name = name
    TemplateColumn.query.filter_by(template_id=template.id).delete()
    for column in columns:
        db.session.add(
            TemplateColumn(
                template_id=template.id,
                column_name=column["name"],
                column_type=column["type"],
                position=column["position"],
            )
        )

    log_event("template_updated", f"Обновлён шаблон '{name}'.")
    db.session.commit()
    return jsonify({"ok": True, "template": serialize_template(template)})


@app.delete("/api/templates/<int:template_id>")
def delete_template(template_id):
    template = Template.query.get_or_404(template_id)
    template_name = template.name
    db.session.delete(template)

    state = get_state_map()
    if state.get("last_template_id") == template_id:
        upsert_state("last_template_id", None)

    log_event("template_deleted", f"Удалён шаблон '{template_name}'.")
    db.session.commit()
    return jsonify({"ok": True})


@app.get("/api/state")
def get_state():
    return jsonify(get_state_map())


@app.put("/api/state")
def save_state():
    payload = request.get_json(silent=True) or {}
    for key in DEFAULT_STATE:
        if key in payload:
            upsert_state(key, payload[key])

    db.session.commit()
    return jsonify({"ok": True, "state": get_state_map()})


@app.get("/api/logs")
def get_logs():
    return jsonify(build_logs_payload())


@app.delete("/api/logs/history")
def clear_logs_history():
    deleted_count = EventLog.query.filter(EventLog.created_at < APP_SESSION_STARTED_AT).delete(
        synchronize_session=False
    )
    log_event("journal_history_cleared", f"История журнала очищена: удалено {deleted_count} записей.")
    db.session.commit()
    return jsonify({"ok": True, "deleted_count": deleted_count, **build_logs_payload()})


@app.get("/api/instructions")
def get_instructions():
    settings = get_settings_map()
    file_path = (settings.get("instructions_file") or "").strip()
    if not file_path:
        return jsonify({"ok": False, "error": "Файл инструкций не задан в настройках."}), 400

    try:
        payload = read_instructions_file(file_path)
    except ValueError as exc:
        log_event("instructions_failed", str(exc))
        db.session.commit()
        return jsonify({"ok": False, "error": str(exc)}), 400

    return jsonify({"ok": True, **payload})


@app.get("/api/data/current")
def get_current_data():
    return jsonify(serialize_dataset())


@app.get("/api/analysis/bootstrap")
def get_analysis_bootstrap():
    dataset = load_analysis_dataset_from_source(get_active_analysis_source())
    CURRENT_ANALYSIS_DATASET.clear()
    CURRENT_ANALYSIS_DATASET.update(dataset)
    return jsonify(
        {
            "types": [serialize_analysis_type(item) for item in AnalysisType.query.order_by(AnalysisType.name).all()],
            "user_state": get_analysis_user_state(),
            "dataset": dataset,
        }
    )


@app.post("/api/analysis/types")
def create_analysis_type():
    payload = request.get_json(silent=True) or {}
    name, charts, errors = validate_analysis_type_payload(payload)
    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    analysis_type = AnalysisType(name=name)
    db.session.add(analysis_type)
    db.session.flush()

    replace_analysis_type_charts(analysis_type, charts)
    upsert_analysis_user_state({"selected_analysis_type_id": analysis_type.id})

    db.session.commit()
    return jsonify({"ok": True, "analysis_type": serialize_analysis_type(analysis_type)}), 201


@app.put("/api/analysis/types/<int:type_id>")
def update_analysis_type(type_id):
    analysis_type = AnalysisType.query.get_or_404(type_id)
    payload = request.get_json(silent=True) or {}
    name, charts, errors = validate_analysis_type_payload(payload, type_id=type_id)
    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    analysis_type.name = name
    replace_analysis_type_charts(analysis_type, charts)
    db.session.commit()
    return jsonify({"ok": True, "analysis_type": serialize_analysis_type(analysis_type)})


@app.delete("/api/analysis/types/<int:type_id>")
def delete_analysis_type(type_id):
    analysis_type = AnalysisType.query.get_or_404(type_id)
    db.session.delete(analysis_type)

    state = AnalysisUserState.query.order_by(AnalysisUserState.id).first()
    if state and state.selected_analysis_type_id == type_id:
        state.selected_analysis_type_id = None

    db.session.commit()
    return jsonify({"ok": True})


@app.put("/api/analysis/state")
def save_analysis_state():
    payload = request.get_json(silent=True) or {}

    next_payload = {}
    if "selected_analysis_type_id" in payload:
        next_payload["selected_analysis_type_id"] = payload.get("selected_analysis_type_id")
    if "left_panel_width" in payload:
        next_payload["left_panel_width"] = int(payload.get("left_panel_width") or DEFAULT_ANALYSIS_USER_STATE["left_panel_width"])
    if "visual_source_kind" in payload:
        value = (payload.get("visual_source_kind") or "none").strip().lower()
        next_payload["visual_source_kind"] = value if value in ALLOWED_ANALYSIS_SOURCES else "none"
    if "table_search" in payload:
        next_payload["table_search"] = str(payload.get("table_search") or "")
    if "table_sort_column" in payload:
        next_payload["table_sort_column"] = str(payload.get("table_sort_column") or "")
    if "table_sort_direction" in payload:
        direction = (payload.get("table_sort_direction") or "asc").strip().lower()
        next_payload["table_sort_direction"] = direction if direction in {"asc", "desc"} else "asc"
    if "draft_state" in payload:
        raw_draft_state = payload.get("draft_state")
        if raw_draft_state:
            analysis_type_id = raw_draft_state.get("analysis_type_id")
            draft_name = str(raw_draft_state.get("name") or "").strip()
            charts, chart_errors = normalize_analysis_charts(raw_draft_state.get("charts"))
            if chart_errors:
                return jsonify({"ok": False, "errors": chart_errors}), 400
            next_payload["draft_state"] = {
                "analysis_type_id": analysis_type_id,
                "name": draft_name,
                "charts": charts,
            }
        else:
            next_payload["draft_state"] = None

    upsert_analysis_user_state(next_payload)
    db.session.commit()
    return jsonify({"ok": True, "user_state": get_analysis_user_state()})


@app.post("/api/analysis/use-facts")
def use_facts_for_analysis():
    try:
        dataset = build_analysis_dataset_from_current_facts_export_layer()
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400

    persist_analysis_source(dataset)
    upsert_analysis_user_state({"visual_source_kind": "facts"})
    log_event("analysis_source_connected", f"Во вкладке 'Анализ' подключён экспортируемый слой из 'Загрузка фактов': '{dataset['source_label']}'.")
    db.session.commit()
    return jsonify({"ok": True, "dataset": serialize_analysis_dataset()})


@app.post("/api/analysis/upload")
def upload_analysis_source():
    file_storage = request.files.get("file")

    try:
        dataset = build_analysis_dataset_from_uploaded_file(file_storage)
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400

    persist_analysis_source(dataset)
    upsert_analysis_user_state({"visual_source_kind": "file"})
    log_event("analysis_file_uploaded", f"Во вкладку 'Анализ' загружен файл '{dataset['source_file_name']}'.")
    db.session.commit()
    return jsonify({"ok": True, "dataset": serialize_analysis_dataset()})


@app.post("/api/analysis/chart-preview")
def get_analysis_chart_preview():
    payload = request.get_json(silent=True) or {}
    chart = payload.get("chart")
    if not isinstance(chart, dict):
        return jsonify({"ok": False, "error": "Некорректная конфигурация графика."}), 400

    preview = build_analysis_chart_preview(chart, serialize_analysis_dataset())
    return jsonify({"ok": True, "preview": preview})


@app.post("/api/analysis/export-table")
def export_analysis_table():
    payload = request.get_json(silent=True) or {}
    raw_columns = payload.get("columns") or []
    raw_rows = payload.get("rows") or []
    file_stem = (payload.get("file_stem") or "analysis_table").strip()
    sheet_title = (payload.get("sheet_title") or "analysis_table").strip()

    if not isinstance(raw_columns, list) or not isinstance(raw_rows, list):
        return jsonify({"ok": False, "error": "Переданы некорректные данные для экспорта таблицы Анализа."}), 400

    columns = []
    seen_columns = set()
    for item in raw_columns:
        name = str(item or "").strip()
        if not name:
            return jsonify({"ok": False, "error": "Список столбцов для экспорта содержит пустое имя."}), 400
        if name.lower() in seen_columns:
            return jsonify({"ok": False, "error": f"Столбец '{name}' передан для экспорта несколько раз."}), 400
        seen_columns.add(name.lower())
        columns.append(name)

    if not columns:
        return jsonify({"ok": False, "error": "Нет столбцов для экспорта таблицы Анализа."}), 400
    if not raw_rows:
        return jsonify({"ok": False, "error": "Нет строк для экспорта таблицы Анализа."}), 400

    output_dir, error = get_output_dir_error()
    if error:
        log_event("analysis_export_failed", error)
        db.session.commit()
        return jsonify({"ok": False, "error": error}), 400

    target_path = build_named_export_path(output_dir, file_stem, "xlsx")

    try:
        save_rows_to_xlsx(target_path, columns, raw_rows, sheet_title=sheet_title)
    except Exception as exc:
        message = f"Не удалось сохранить таблицу Анализа в XLSX: {exc}"
        log_event("analysis_export_failed", message)
        db.session.commit()
        return jsonify({"ok": False, "error": message}), 500

    log_event(
        "analysis_export_success",
        f"Таблица Анализа сохранена в файл '{target_path.name}': {len(raw_rows)} строк, {len(columns)} столбцов.",
    )
    db.session.commit()

    return jsonify(
        {
            "ok": True,
            "file_name": target_path.name,
            "file_path": str(target_path),
            "row_count": len(raw_rows),
            "column_count": len(columns),
        }
    )


@app.post("/api/data/upload")
def upload_data():
    global LAST_PARSE_SNAPSHOT

    file_storage = request.files.get("file")
    template_id = request.form.get("template_id", type=int)
    template = Template.query.get(template_id) if template_id else None

    try:
        dataset = build_dataset(file_storage, template)
    except ValueError as exc:
        log_event("upload_failed", str(exc))
        db.session.commit()
        return jsonify({"ok": False, "error": str(exc)}), 400

    CURRENT_DATASET.clear()
    CURRENT_DATASET.update(dataset)
    LAST_PARSE_SNAPSHOT = None

    error_count = len(dataset["errors"])
    warning_count = len(dataset["warnings"])
    if dataset["partial"]:
        event_type = "upload_partial"
        description = (
            f"Файл '{dataset['file_name']}' загружен частично: "
            f"{dataset['row_count']} строк, {warning_count} предупреждений, {error_count} ошибок."
        )
    else:
        event_type = "upload_success"
        description = (
            f"Файл '{dataset['file_name']}' загружен: "
            f"{dataset['row_count']} строк, {warning_count} предупреждений, {error_count} ошибок."
        )

    log_event(event_type, description)
    log_dataset_quality(dataset, f"Загрузка файла '{dataset['file_name']}'")
    db.session.commit()

    return jsonify({"ok": True, "dataset": dataset})


@app.post("/api/data/parse")
def parse_data():
    global LAST_PARSE_SNAPSHOT

    if not CURRENT_DATASET["rows"]:
        return jsonify({"ok": False, "error": "Нет данных для парсинга."}), 400

    payload = request.get_json(silent=True) or {}
    column_name = (payload.get("column_name") or "").strip()
    delimiter = str(payload.get("delimiter") or "")
    row_ids = payload.get("row_ids") or []

    if not column_name:
        return jsonify({"ok": False, "error": "Выберите столбец для парсинга."}), 400
    if not delimiter:
        return jsonify({"ok": False, "error": "Укажите разделитель парсинга."}), 400
    if column_name not in {column["name"] for column in CURRENT_DATASET["columns"]}:
        return jsonify({"ok": False, "error": "Выбранный столбец не найден в рабочем наборе."}), 400

    try:
        affected_row_ids = {int(item) for item in row_ids}
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Передан некорректный набор строк для парсинга."}), 400

    if not affected_row_ids:
        return jsonify({"ok": False, "error": "Нет строк после фильтрации для парсинга."}), 400

    LAST_PARSE_SNAPSHOT = deepcopy(CURRENT_DATASET)

    deleted_rows = 0
    created_rows = 0
    next_row_id = CURRENT_DATASET.get("next_row_id", 1)
    updated_rows = []

    for row in CURRENT_DATASET["rows"]:
        if row.get("row_id") not in affected_row_ids:
            updated_rows.append(row)
            continue

        deleted_rows += 1
        fragments = split_parse_fragments(row["values"].get(column_name, ""), delimiter)
        if not fragments:
            continue

        for fragment in fragments:
            new_values = dict(row["values"])
            new_values[column_name] = fragment
            updated_rows.append(
                {
                    "row_id": next_row_id,
                    "row_number": 0,
                    "source_row_number": row["source_row_number"],
                    "values": new_values,
                }
            )
            next_row_id += 1
            created_rows += 1

    rebuild_row_numbers(updated_rows)
    CURRENT_DATASET["rows"] = updated_rows
    CURRENT_DATASET["row_count"] = len(updated_rows)
    CURRENT_DATASET["errors"] = build_dataset_errors(updated_rows, CURRENT_DATASET["columns"])
    CURRENT_DATASET["next_row_id"] = next_row_id
    CURRENT_DATASET["can_undo_parse"] = True

    log_event(
        "parse_applied",
        f"Парсинг по столбцу '{column_name}' выполнен: удалено {deleted_rows}, создано {created_rows}.",
    )
    log_dataset_quality(serialize_dataset(), f"Парсинг по столбцу '{column_name}'")
    db.session.commit()

    return jsonify(
        {
            "ok": True,
            "dataset": serialize_dataset(),
            "summary": {
                "deleted_rows": deleted_rows,
                "created_rows": created_rows,
            },
        }
    )


@app.post("/api/data/parse/undo")
def undo_parse():
    global LAST_PARSE_SNAPSHOT

    if LAST_PARSE_SNAPSHOT is None:
        return jsonify({"ok": False, "error": "Нет доступного шага отката парсинга."}), 400

    restored_dataset = deepcopy(LAST_PARSE_SNAPSHOT)
    restored_dataset["can_undo_parse"] = False

    CURRENT_DATASET.clear()
    CURRENT_DATASET.update(restored_dataset)
    LAST_PARSE_SNAPSHOT = None

    log_event("parse_undo", "Выполнен откат последнего парсинга.")
    db.session.commit()

    return jsonify({"ok": True, "dataset": serialize_dataset()})


@app.post("/api/data/export")
def export_data():
    if not CURRENT_DATASET["columns"]:
        return jsonify({"ok": False, "error": "Нет данных для сохранения."}), 400

    payload = request.get_json(silent=True) or {}
    raw_columns = payload.get("columns") or []
    raw_rows = payload.get("rows") or []

    if not isinstance(raw_columns, list) or not isinstance(raw_rows, list):
        return jsonify({"ok": False, "error": "Переданы некорректные данные для сохранения."}), 400

    columns = []
    seen_columns = set()
    for item in raw_columns:
        name = (item.get("name") or "").strip()
        if not name:
            return jsonify({"ok": False, "error": "Список столбцов для сохранения содержит пустое имя."}), 400
        if name.lower() in seen_columns:
            return jsonify({"ok": False, "error": f"Столбец '{name}' передан для сохранения несколько раз."}), 400
        seen_columns.add(name.lower())
        columns.append(name)

    if not columns:
        return jsonify({"ok": False, "error": "Нет видимых столбцов для сохранения."}), 400

    settings = get_settings_map()
    output_folder = (settings.get("output_folder") or "").strip()
    if not output_folder:
        message = "Не задана папка результата в настройках."
        log_event("export_failed", message)
        db.session.commit()
        return jsonify({"ok": False, "error": message}), 400

    output_dir = Path(output_folder)
    if not output_dir.exists() or not output_dir.is_dir():
        message = "Папка результата недоступна. Проверьте настройки приложения."
        log_event("export_failed", message)
        db.session.commit()
        return jsonify({"ok": False, "error": message}), 400

    target_path = build_export_path(output_dir)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "data"
    worksheet.append(columns)

    try:
        for row in raw_rows:
            worksheet.append([str((row or {}).get(column_name, "")) for column_name in columns])
        workbook.save(target_path)
    except Exception as exc:
        message = f"Не удалось сохранить XLSX: {exc}"
        log_event("export_failed", message)
        db.session.commit()
        return jsonify({"ok": False, "error": message}), 500

    log_event(
        "export_success",
        f"Результат сохранён в файл '{target_path.name}': {len(raw_rows)} строк, {len(columns)} столбцов.",
    )
    db.session.commit()

    return jsonify(
        {
            "ok": True,
            "file_name": target_path.name,
            "file_path": str(target_path),
            "row_count": len(raw_rows),
            "column_count": len(columns),
        }
    )


with app.app_context():
    db.create_all()
    ensure_analysis_chart_schema()
    ensure_analysis_user_state_schema()


if __name__ == "__main__":
    app.run(debug=True)